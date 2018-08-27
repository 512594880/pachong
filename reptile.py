import asyncio
import aiohttp
import time
import os
import xlwt
import xlrd
from xlutils.copy import copy
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook

sheetsize = 0
row = 1
wbk = xlwt.Workbook(encoding = 'utf-8')
sheet = wbk.add_sheet('sheet 1')

def saveExcel(str,row,colum):
    global sheetsize
    sheet = wbk.get_sheet(sheetsize)
    sheet.write(row, colum,  str)
    return

def saveInExcel(name,str,row,colum):

    excelName = '/Users/wangxi/Desktop/'+name +'.xls'
    path = os.path.join(excelName)
    exists = os.path.exists(path)

    if exists:
        print("excel 存在")
        wbk = xlrd.open_workbook(excelName,formatting_info=True)
        wbknew = copy(wbk)
        wbknew.encoding = 'utf-8'
        wbk = wbknew
        sheet = wbk.get_sheet(0)
    else:
        print("excel不存在")
        wbk = xlwt.Workbook(encoding = 'utf-8')
        sheet = wbk.add_sheet('sheet 1')
        sheet = wbk.get_sheet(0)
    sheet.write(row, colum,  str)
    wbk.save(excelName)
    return


def saveInExcelWithXlsx(name,str,row,colum):
    excelName = '/Users/wangxi/Desktop/'+name +'.xlsx'
    path = os.path.join(excelName)
    exists = os.path.exists(path)

    if exists:
        print("excel 存在")
        wbk = load_workbook(excelName)
        wbknew = copy(wbk)
        wbknew.encoding = 'utf-8'
        wbk = wbknew
        sheet = wbk.get_sheet(0)

    else:
        print("excel不存在")
        wbk = openpyxl.Workbook()
        size = wbk.get_sheet_names().count()
        sheet = wbk.create_sheet(size)
    sheet.write(row, colum,  str)
    wbk.save(excelName)
    return






def handlResult(result):
    print("---------")
    colum = 0
    global row
    soup = BeautifulSoup(result)
    liResutl = soup.findAll('tbody')
    for li in liResutl:
        trlist = li.findAll('tr')
        for tr in trlist:
            # valuelist = []
            tdlist = tr.findAll('td')
            thlist = tr.findAll('th')
            for th in thlist:
                value = ""
                if 'href' in th:
                    name = th.find('a',attrs = {"class" : "cl-blue"})
                    value= name.get_text()
                else:
                    value = th.get_text()
                # valuelist.append(value)
                saveExcel(value,row,colum)
                colum = colum +1
            for td in tdlist:
                value = ""
                if 'href' in td:
                    name = td.find('a',attrs = {"class" : "cl-blue"})
                    value= name.get_text()
                else:
                    value = td.get_text()
                # valuelist.append(value)
                saveExcel(value,row,colum)
                colum = colum +1
            colum = 0
            row = row +1
            # valueArray.append(valuelist)
            print("----" + str(row) +"--------")

async def get(url):

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}

    session = aiohttp.ClientSession(headers=headers)

    response = await session.get(url)

    result = await response.text()

    session.close()

    return result





async def requestBynum(i,name):

    url = 'https://db.yaozh.com/qixiebiaozhun?p='+str(i+1)+'&pageSize=20'

    print('Waiting for',url)

    result = await get(url)

    if name == 'qixiebiaozhun':
        handlQixiebiaozhun(result,name)
    elif name == "zhuce":
        handlResult(result)

def handlQixiebiaozhun(result,name):
	print("---------")
	colum = 0
	global row
	soup = BeautifulSoup(result)
	liResutl = soup.findAll('tbody')
	for li in liResutl:
		trlist = li.findAll('tr')
		for tr in trlist:
			print(tr)



if __name__ == '__main__':
    excelName = '/Users/anne/Desktop/' + "药品注册与受理" + '.xls'
    start = time.time()
    for y in range(1):
        tasks = [asyncio.ensure_future(requestBynum(i,"qixiebiaozhun")) for i in range(y, y + 1)]
        loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))

    end = time.time()

    print('Cost time:', end - start)
